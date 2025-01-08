
# -*- coding: utf-8 -*-
"""
Created on Wed Dec 18 10:49:58 2024

@author: sophi

source documentation : https://dash-bootstrap-components.opensource.faculty.ai/docs/components/accordion/
http://127.0.0.1:8050

GET LOCAL IP 
# Trouver l'adresse IP locale
import socket
hostname = socket.gethostname()
local_ip = socket.gethostbyname(hostname)
print(f"Adresse IP locale : {local_ip}")
"""
#cd C:\Users\sophi\myCloud\Sophia\Thesis\Model\Jucar_model\Adrià\App_Jucar
#streamlit run Jucar_river_basin.py
#git save

# Import libraries and modules
import pandas as pd
import numpy as np
import openpyxl
import pysd
import dash
from dash import Dash, dcc, html, Input, Output, State
import dash_bootstrap_components as dbc
import plotly.graph_objs as go
import boto3



def upload_to_s3(file_name, bucket_name, object_name=None):
    # S3 client
    s3_client = boto3.client('s3', aws_access_key_id='YOUR_ACCESS_KEY', aws_secret_access_key='YOUR_SECRET_KEY')

    if object_name is None:
        object_name = file_name

    try:
        s3_client.upload_file(file_name, bucket_name, object_name)
        print(f"Fichier {file_name} sauvegardé dans le bucket S3 : {bucket_name}")
    except Exception as e:
        print(f"Erreur lors de l'envoi à S3 : {e}")



# Precompute initial graph data
workbook = openpyxl.load_workbook("data_initial.xlsx")
workbook.save("data.xlsx")
upload_to_s3("data.xlsx", "mon-bucket-s3")
vensim_model = pysd.load('WEFE Jucar (Simple).py')
years_sim = 10
months = np.arange(1, (12*years_sim)+1)
variables_model_initial = vensim_model.run(params={'INITIAL TIME': 1, 'FINAL TIME': 12*years_sim, 'TIME STEP': 1})
#1.Alarcon initial values
initial_qecolAlar_value = 5.18
qecolAlar_value = initial_qecolAlar_value
initial_outflow = variables_model_initial['Sal Jucar']
initial_deficit = variables_model_initial['DéfQecolAlar']
#1.Population growth initial value
initial_variation_rate = 0.00018728
variation_rate = initial_variation_rate
initial_urban_demand = variables_model_initial['Total Demanda Urbana']
# Generate labels for months and years
months_labels = []

# Initialize the Dash app
app = Dash(__name__, external_stylesheets=[dbc.themes.FLATLY, "https://cdn.jsdelivr.net/npm/bootstrap-icons/font/bootstrap-icons.css"], suppress_callback_exceptions=True)
app.title = "Júcar River Basin Water Management"
server = app.server
# Define reusable page components
#### 1. Home page
def create_home_page():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        html.P("Welcome to the Júcar River Basin System Dynamics Model. This tool provides an interactive platform "
               "to analyze and simulate the behavior of the Júcar River Basin under various scenarios."),
        html.H2("Content", className="mt-4"),
        html.Ul([
            html.Li("1. Model presentation - see how the model is formed"),
            html.Li("2. Alarcón’s Reservoir - Environmental flow simulation"),
            html.Li("3. Population Growth Analysis"),
            html.Li("4. Diverse Graphs"),
            html.Li("jgfodijgoidjfgoijdfgjfoigjofjfjg"),
        ]),
        html.H2("User Guide", className="mt-4"),
        html.Ol([
            html.Li("Navigate using the menu on the left."),
            html.Li("Adjust variables using sliders or select different scenarios using dropdown menus."),
            html.Li("Run simulations to analyze results."),
        ])
    ])


#### 2. Model presentation page
def create_model_presentation():
    return html.Div([
        html.H1("Júcar River Basin Management Tool", className="text-center mt-0"),
        dbc.Tabs(
            [
                dbc.Tab(label="View 1: SYSTEM NETWORK", children=[
                    html.P("Overview of the model"),
                    html.Img(src="/assets/View_1.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Aquifer", children=[
                    html.P("View 2: MANCHA ORIENTAL AQUIFER"),
                    html.Img(src="/assets/View_2.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Water demand, supply and deficit", children=[
                    html.P("View 3: WATER DEMAND, SUPPLY AND DEFICIT"),
                    html.Img(src="/assets/View_3.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Reservoirs operating rules", children=[
                    html.P("View 4: RESERVOIRS OPERATING RULES"),
                    html.Img(src="/assets/View_4.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="State index", children=[
                    html.P("View 5: STATE INDEX"),
                    html.Img(src="/assets/View_5.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
                dbc.Tab(label="Crops", children=[
                    html.P("View : CROPS"),
                    html.Img(src="/assets/Crops.PNG", style={"width": "100%", "margin": "auto", "display": "block"})
                ]),
            ]
        )
    ])

#### 3. Alarcon page
## 3.1 Modal : parameters explanation
def create_modal_Alarcon_Qeco():
    modal_Alarcon = html.Div([
            html.Div([
                    html.P(
                        "Parameter : QEcolAlar",
                        style={"display": "inline-block","margin-right": "5px","text-align": "center"},
                    ),
                    html.I(
                        className="bi bi-info-circle",
                        id="info-icon-Alarcon-Qeco",  # This ID is used to open the modal
                        style={"fontSize": "0.9rem","color": "#0f375e","cursor": "pointer",},
                    ),
                ],
                style={"display": "flex", "align-items": "center","margin-top": "20px","margin-top": "20px","margin-bottom": "10px"},
            ),
            # Modal
            dbc.Modal([
                    dbc.ModalHeader(dbc.ModalTitle("Explanation")),
                    dbc.ModalBody(
                        html.Div(
                            [
                                html.P(
                                    "QEcolAlar: This is the environmental flow downstream of Alarcón's reservoir. "
                                    "It is based on hydrological, operational, and environmental factors."
                                    "This is the Causal diagram :"
                                ),
                                html.Img(src="/assets/QecoAlar_tree.jpg", style={"width": "100%", "margin-bottom": "20px"}, alt="Diagram of QEcolAlar parameter",),
                                html.P(
                                    "In red is it the parameter Qeco in the System Dynamic Model "
                                    "In green are the two outputs of the simulation (see graph)"
                                    "Sal Jucar : the ouflow "
                                    "DéfQecolAlar : the deficit regarding the environmental flow " 
                                ),
                                html.Img(src="/assets/Output.jpg", style={"width": "100%", "margin-bottom": "20px"}, alt="Diagram of QEcolAlar parameter",),
                            ]
                        ),
                    ),
                    dbc.ModalFooter(
                        dbc.Button(
                            "Close", id="close-Alarcon-Qeco", className="ms-auto", n_clicks=0
                        )
                    ),
                ],
                id="modal-Alarcon-Qeco",
                is_open=False,
            ),
        ]
    )
    return modal_Alarcon

def create_modal_Alarcon_inputs():
    modal_Alarcon = html.Div([
            html.Div([
                    html.H5(
                        "Select input type ",
                        style={"display": "inline-block","margin-right": "5px","margin-top": "5px"},
                    ),
                    html.I(
                        className="bi bi-info-circle",
                        id="info-icon-Alarcon-inputs",  # ID  to open the modal
                        style={"fontSize": "0.7rem","color": "#0f375e","cursor": "pointer",},
                    ),
                    html.P(
                        ":",
                        style={"display": "inline-block","margin-left": "5px"},
                    ),
                ],style={"display": "flex", "align-items": "center"}, 
            ),
            # Modal
            dbc.Modal([
                    dbc.ModalHeader(dbc.ModalTitle("Explanation of Parameters")),
                    dbc.ModalBody(
                        html.Div([
                                html.P([html.B("Constant"), " : means QEcolAlar is constant over the months."]),
                                html.P([html.B("Montly dynamic"), " : means QEcolAlar can be modified for more flexibility for each month."]),
                            ]
                        )
                    ),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-Alarcon-inputs", className="ms-auto", n_clicks=0)
                    ),
                ],
                id="modal-Alarcon-inputs",
                is_open=False,
            ),
        ]
    )
    return modal_Alarcon


# 3.2 Parameter panel
def create_parameter_panel_Alarcon():
    return html.Div([
        html.H3("Parameter Settings", className="text-center mt-3"),
        create_modal_Alarcon_Qeco(),
        # Dropdowns for year range selection
        html.Div([
            html.Div([
                html.P("Select simulation range:", className="text-center mt-1"),
                dbc.Row([
                    dbc.Col([
                        html.Label("Start Year"),
                        dcc.Dropdown(id="start-year",options=[{"label": str(year), "value": year} for year in range(2003, 2014)],value=2003,clearable=False,style={"width": "100%"}) ], width=6),
                    dbc.Col([
                        html.Label("End Year"),
                        dcc.Dropdown(id="end-year",options=[{"label": str(year), "value": year} for year in range(2003,2014)], value=2013,clearable=False,style={"width": "100%"})], width=6),
                ], className="mb-3")
            ], style={"margin-bottom": "20px"}),

            # choose between constant or dynamic dropdown
            create_modal_Alarcon_inputs(),
            dbc.RadioItems(
                id="qecolAlar-selector",
                options=[
                    {"label": "Constant", "value": "option1"},
                    {"label": "Monthly dynamic", "value": "option2"},
                ],
                inline=True,
                value="option1",# Default selection
                style={"margin-left": "18px"},
                className="mb-3"
            ),
            html.Hr(style={"border": "1px solid black","margin-bottom": "20px","margin-top": "25px"}), #black horizontal line
            
            # Option 1 : Dropdown for constant input 
            html.Div([
                html.P("Set a constant Qeco:", className="text-center mt-1"),
                dbc.Input(id="qecolAlar-constant-input",type="number", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value,
                          style={"width": "100px", "text-align": "center", "margin": "0 auto"})
            ], id="qecolAlar-constant", style={"text-align": "center", "margin-bottom": "20px"}),

            # Option 2 : Dynamic dropdown for 12 months for dynamic input
            html.Div([
                html.P("Set a Qeco for each month:", className="text-center mt-1"),
                html.Div([
                    dbc.Row([
                        dbc.Col([html.Label("January"), dbc.Input(type="number", id="jan-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("February"), dbc.Input(type="number", id="feb-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ], className="mb-2"),
                    dbc.Row([
                        dbc.Col([html.Label("March"), dbc.Input(type="number", id="mar-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("April"), dbc.Input(type="number", id="apr-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ], className="mb-2"),
                    dbc.Row([
                        dbc.Col([html.Label("May"), dbc.Input(type="number", id="may-value",min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("June"), dbc.Input(type="number", id="jun-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ], className="mb-2"),
                    dbc.Row([
                        dbc.Col([html.Label("July"), dbc.Input(type="number", id="jul-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("August"), dbc.Input(type="number", id="aug-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ], className="mb-2"),
                    dbc.Row([
                        dbc.Col([html.Label("September"), dbc.Input(type="number", id="sep-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("October"), dbc.Input(type="number", id="octo-value",min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ], className="mb-2"),
                    dbc.Row([
                        dbc.Col([html.Label("November"), dbc.Input(type="number", id="nov-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                        dbc.Col([html.Label("December"), dbc.Input(type="number", id="dec-value", min=0.0, max=10.0, step=0.1, value=initial_qecolAlar_value)], width=6),
                    ])
                ])
            ], id="qecolAlar-dynamic", 
                style={"display": "none", "margin-bottom": "20px","margin-top": "20px"},
                className="text-center mt-0"),
            
            #simulation button
            html.Div([
                dbc.Button("Run Simulation", id="run-simulation", color="primary",className="text-center mt-0")
            ],  style={"display": "flex", "justify-content": "center","align-items": "center", "margin-top": "20px",  "margin-bottom": "20px"})
        ], className="p-4 bg-light shadow rounded", style={ "margin-bottom": "20px"}),
        
    ], style={"maxWidth": "600px", "margin": "20px auto","boxShadow": "0 4px 8px rgba(0,0,0,0.1)","backgroundColor": "#dbe1e7","borderRadius": "8px","padding": "25px"})


def create_alarcon_page():
    return dbc.Container(fluid=True,children=[
        html.H1("Alarcón’s Reservoir", className="text-center mt-0"),
        dbc.Row([
            dbc.Col(create_parameter_panel_Alarcon(), width=3),
            dbc.Col([
                dbc.Spinner(
                html.Div([
                    dcc.Graph(
                        id="outflow-graph",
                        figure={
                            "data": [go.Scatter(x=months, y=initial_outflow, mode="lines", name="Outflow")],
                            "layout": go.Layout(title="Outflow Over Time", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                        },style={"border": "0.5px solid rgba(15, 55, 94, 0.3)","padding": "15px", "box-shadow": "0 4px 8px rgba(0, 0, 0, 0.1)","background-color": "#ffffff",}),
                    dcc.Graph(
                        id="deficit-graph",
                        figure={
                            "data": [go.Scatter(x=months, y=initial_deficit, mode="lines", name="Deficit")],
                            "layout": go.Layout(title="Deficit Over Time (DéfQEcolAlar)", xaxis={"title": "Months"}, yaxis={"title": "hm³"})
                        },style={"border": "0.5px solid rgba(15, 55, 94, 0.3)","padding": "15px", "box-shadow": "0 4px 8px rgba(0, 0, 0, 0.1)","background-color": "#ffffff",}),
                ],),)
                
            ], width=9),
        ])
    ])


#### 4. Population growth simulation page
## 4.1 Modal : parameters explanation
def create_modal_population():
    modal_population = html.Div([
            html.Div([
                    html.P(
                        "Parameter : Variation rate",
                        style={"display": "inline-block","margin-right": "5px","text-align": "center"},
                    ),
                    html.I(
                        className="bi bi-info-circle",
                        id="info-icon-population",  # This ID is used to open the modal
                        style={"fontSize": "0.9rem","color": "#0f375e","cursor": "pointer",},
                    ),
                ],style={"display": "flex"}, 
            ),
            # Modal
            dbc.Modal([
                    dbc.ModalHeader(dbc.ModalTitle("Explanation")),
                    dbc.ModalBody(
                        html.Div([
                                html.P(
                                    "Variation rate: Simulate the population growth for Valencia and Sagunto."
                                    "This is the Causal diagrams :"
                                    ),
                                html.Img(src="/assets/links_variation_rate.png", style={"width": "100%", "margin-bottom": "20px"}, alt="Diagram of Variation rate parameter",),
                            ]
                        ),
                    ),
                    dbc.ModalFooter(
                        dbc.Button("Close", id="close-population", className="ms-auto", n_clicks=0)
                                    ),
                ],
                id="modal-population",
                is_open=False,
            ),
        ]
    )
    return modal_population


## 4.2 parameter panel
def create_parameter_panel_population():
    return html.Div([
        html.H3("Parameter Settings", className="text-center mt-3"),
        create_modal_population(),
        # Dropdowns for year range selection
        html.Div([
            html.P("Select simulation range:", className="text-center mt-1"),
            dbc.Row([
                dbc.Col([
                    html.Label("Start Year"),
                    dcc.Dropdown(id="start-year", options=[{"label": str(year), "value": year} for year in range(2003, 2014)], value=2003, clearable=False, style={"width": "100%"} )], width=6),
                dbc.Col([
                    html.Label("End Year"),
                    dcc.Dropdown(id="end-year", options=[{"label": str(year), "value": year} for year in range(2003, 2014)], value=2013,  clearable=False,  style={"width": "100%"} )], width=6),
            ], className="mb-3")
        ], style={"margin-bottom": "20px"}),
        html.Div([
        #Dropdown  for variation rate input
        html.Div([
                html.P("Set Variation rate:", className="text-center mt-1"),
                dbc.Input(id="variation-rate-input",type="number", min=0.000, max=1.000, step=0.0001, value=initial_variation_rate,
                          style={"width": "100px", "text-align": "center", "margin": "0 auto"})
                ], id="variation-rate", style={"text-align": "center", "margin-bottom": "10px"}),
         #simulation button
                html.Div([
                    dbc.Button("Run Simulation", id="run-simulation", color="primary",className="text-center mt-0")
                ],  style={"display": "flex", "justify-content": "center","align-items": "center", "margin-top": "20px",  "margin-bottom": "20px"})
                ],className="p-4 bg-light shadow rounded", style={ "margin-bottom": "20px"})
        ], style={"maxWidth": "600px", "margin": "20px auto", "boxShadow": "0 4px 8px rgba(0,0,0,0.1)","backgroundColor": "#dbe1e7","borderRadius": "8px","padding": "25px"})




def create_population_growth_page():
    return dbc.Container(fluid=True,children=[
            html.H1("Population Growth Analysis", className="text-center mt-0"),
            dbc.Row([
                    dbc.Col(create_parameter_panel_population(), width=3),
                    dbc.Col([
                            dbc.Spinner(
                                html.Div([
                                        html.P("Population Growth Dynamics: Analyze how population growth impacts urban demand."),
                                        html.Div([
                                                dcc.Graph(id="demand-graph",
                                                    figure={
                                                        "data": [go.Scatter(x=months, y=initial_urban_demand,mode="lines",name="Total Demanda Urbana",)],
                                                        "layout": go.Layout( title="Total Demanda Urbana", xaxis={"title": "Months"},yaxis={"title": "hm³"},legend={"orientation": "h", "x": 0,"y": 1.2, "xanchor": "left", "yanchor": "bottom", },margin={"l": 40, "r": 40, "t": 40, "b": 40},),  # Graph margin
                                                             },
                                                        ),
                                                ],style={"border": "0.5px solid rgba(15, 55, 94, 0.3)","padding": "15px", "box-shadow": "0 4px 8px rgba(0, 0, 0, 0.1)","background-color": "#ffffff",},),
                                        ])
                                        )
                            ], width=9,),
                    ]),
            ],)
### 5. Diverse Graphs pages
## 5.1 parameter panel
def create_parameter_panel_divers():
    variable_headers = list(variables_model_initial.columns)  
    return html.Div([
        html.H3("Parameter Settings", className="text-center mt-3"),
        # Dropdowns for year range selection
        html.Div([
            html.P("Select simulation range:", className="text-center mt-1"),
            dbc.Row([
                dbc.Col([
                    html.Label("Start Year"),
                    dcc.Dropdown(id="start-year",options=[{"label": str(year), "value": year} for year in range(2003, 2014)],value=2003,clearable=False,style={"width": "100%"}  )], width=6),
                dbc.Col([
                    html.Label("End Year"),
                    dcc.Dropdown(id="end-year",options=[{"label": str(year), "value": year} for year in range(2003, 2014)],value=2013,clearable=False,style={"width": "100%"} )], width=6),
            ], className="mb-3")
        ], style={"margin-bottom": "20px"}),
        # Dropdown to select a variable
        html.Div([
            html.P("Select a graph variable:", className="text-center mt-1"),
            dbc.Select(id="selected",options=[{"label": header, "value": header} for header in variable_headers],placeholder="Select a graph variable", style={"width": "100%", "text-align": "center", "margin": "0 auto"} )
        ], style={"text-align": "center", "margin-bottom": "10px"}),

        # Simulation button
        html.Div([
            dbc.Button("Run Simulation", id="run-simulation", color="primary", className="text-center mt-0")
        ], style={"display": "flex", "justify-content": "center", "align-items": "center", "margin-top": "20px", "margin-bottom": "20px"})
    ], style={"maxWidth": "600px", "margin": "20px auto", "boxShadow": "0 4px 8px rgba(0,0,0,0.1)", "backgroundColor": "#dbe1e7", "borderRadius": "8px", "padding": "25px"})

def create_divers_graphs_page():
    return dbc.Container(fluid=True,children=[
            html.H1("Diverse Graphs", className="text-center mt-0"),
            dbc.Row([
                    dbc.Col(create_parameter_panel_divers(), width=3),
                    dbc.Col([
                            dbc.Spinner(
                                html.Div([
                                        html.Div([
                                                dcc.Graph(id="divers-graph", figure={},),
                                                ],style={"border": "0.5px solid rgba(15, 55, 94, 0.3)","padding": "15px", "box-shadow": "0 4px 8px rgba(0, 0, 0, 0.1)","background-color": "#ffffff",},),
                                        ])
                                        )], width=9,),]),],)

# Navigation menu
def create_menu():
    return dbc.Nav(
        [
            dbc.NavLink("Home", href="/", active="exact", className="dropdown-item"),
            dbc.NavLink("Model presentation", href="/model", active="exact", className="dropdown-item"),
            dbc.NavLink("Alarcón’s Reservoir", href="/alarcon", active="exact", className="dropdown-item"),
            dbc.NavLink("Population Growth", href="/population-growth", active="exact", className="dropdown-item"),
            dbc.NavLink("Diverse Graphs", href="/diverse-graphs", active="exact", className="dropdown-item"),
        ],
        pills=True,
        vertical=True,
        className="navbar",
        style={"width": "250px","position": "fixed","left": "0","top": "0","background-color": "#f8f9fa","padding": "45px 10px","box-shadow": "2px 0 5px rgba(0,0,0,0.1)",},
    )


# Update page routing
@app.callback(
    Output("page-content", "children"),
    Input("url", "pathname")
)
def update_page(pathname):
    if pathname == "/":
        return create_home_page()
    elif pathname == "/model":
        return create_model_presentation()
    elif pathname == "/alarcon":
        return create_alarcon_page()
    elif pathname == "/population-growth":
        return create_population_growth_page()
    elif pathname == "/diverse-graphs":
        return create_divers_graphs_page()
    return html.Div("404: Page Not Found")
# Layout
app.layout = html.Div([
    dbc.Button("Menu", id="menu-toggle", color="primary", className="mb-2", style={"position": "fixed", "top": "10px", "left": "10px", "zIndex": 1000}),
    dbc.Collapse(create_menu(), id="menu-collapse", is_open=False),
    html.Div([
        dcc.Location(id="url"),
        html.Div(id="page-content", style={"padding": "20px"})
    ], id="main-content", style={"margin-left": "0px", "transition": "margin-left 0.3s ease"})
])


@app.callback(
    [Output("menu-collapse", "is_open"),# Control the is_open state of the Collapse
     Output("main-content", "style")],
    [Input("menu-toggle", "n_clicks")],# Triggered when the toggle button is clicked
    [State("menu-collapse", "is_open")], # Store the current state of the Collapse
)
def toggle_menu(n_clicks, is_open):
    if n_clicks:
        if not is_open:
            return True, {"margin-left": "250px", "transition": "margin-left 0.3s ease"}
        return False, {"margin-left": "0px", "transition": "margin-left 0.3s ease"}
    return is_open, {"margin-left": "0px", "transition": "margin-left 0.3s ease"}



### 3.4 Callback Alarcon
#modal Qec0
@app.callback(
    Output("modal-Alarcon-Qeco", "is_open"),
    [Input("info-icon-Alarcon-Qeco", "n_clicks"), Input("close-Alarcon-Qeco", "n_clicks")],
    [State("modal-Alarcon-Qeco", "is_open")],
)
def toggle_modal_Alarcon_Qeco(icon_click, close_click, is_open):
    if icon_click or close_click:
        return not is_open
    return is_open

#modal inputs
@app.callback(
    Output("modal-Alarcon-inputs", "is_open"),
    [Input("info-icon-Alarcon-inputs", "n_clicks"), Input("close-Alarcon-inputs", "n_clicks")],
    [State("modal-Alarcon-inputs", "is_open")],
)
def toggle_modal_Alarcon_inputs(icon_click, close_click, is_open):
    if icon_click or close_click:
        return not is_open
    return is_open


#selection of dropdown value
@app.callback(
    [Output("qecolAlar-constant", "style"),
     Output("qecolAlar-dynamic", "style"),
     Output("run-simulation", "style")],
    Input("qecolAlar-selector", "value")
)

def toggle_input(input_type):
    if input_type == "option1":
        # Show dropdown constant, hide dropdown monthly
        return {"display": "block"}, {"display": "none"}, {"display": "block", "margin-bottom": "20px"}
    elif input_type == "option2":
        # Show dropdown, hide slider
        return {"display": "none"}, {"display": "block", "margin-bottom": "20px","margin-top": "20px"} ,{"display": "block", "margin-bottom": "20px"}
    # If no option is selected, hide both
    return {"display": "none"}, {"display": "none"},{"display": "none"}

#graphs
# Graph callbacks
@app.callback(
    [Output("outflow-graph", "figure"),
     Output("deficit-graph", "figure")],
    [Input("run-simulation", "n_clicks")],
    [State("start-year", "value"),
     State("end-year", "value"),
     State("qecolAlar-selector", "value"),
     State("qecolAlar-constant-input", "value"),
     State("jan-value", "value"),
     State("feb-value", "value"),
     State("mar-value", "value"),
     State("apr-value", "value"),
     State("may-value", "value"),
     State("jun-value", "value"),
     State("jul-value", "value"),
     State("aug-value", "value"),
     State("sep-value", "value"),
     State("octo-value", "value"),
     State("nov-value", "value"),
     State("dec-value", "value")],
    prevent_initial_call=True
)
def update_Alarcon_graphs(n_clicks, start_year, end_year, input_type, constant_value, jan, feb, mar, apr, may, jun, jul, aug, sep, octo, nov, dec):
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate  # Prevent callback if no clicks

    
    # Calculate the number of months based on the selected years
    years_sim = end_year - start_year + 1
    months = np.arange(1, 12 * years_sim + 1)
    
    # Calculate the number of months based on the selected years
    years_sim = end_year - start_year + 1
    months = np.arange(1, 12 * years_sim + 1)
    # Generate labels for all months and October-only
    all_months_labels = []
    october_labels = []
    tickvals_october = []
    for year_index, year in enumerate(range(start_year, end_year + 1)):
        for month_index, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
            all_months_labels.append(f"{month} {year}")
            if month == "Oct":
                tickvals_october.append(year_index * 12 + month_index + 1)
                october_labels.append(f"Oct {year}")
    # Determine tick labels based on simulation length
    if years_sim > 3:
        tickvals = tickvals_october
        ticktext = october_labels
    else:
        tickvals = list(range(1, len(all_months_labels) + 1))
        ticktext = all_months_labels

    # Determine QecoAlar values based on input type
    if input_type == "option1":
        qecolAlar_values = [constant_value] * 12
    else:
        qecolAlar_values = [jan, feb, mar, apr, may, jun, jul, aug, sep, octo, nov, dec]

    # Update the Excel file
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook["Demandas"]
    column_name = "QecolAlar"

    # Find the column index for QecolAlar
    column_index = None
    for col in sheet[2]:  # Assuming header is in row 2
        if col.value == column_name:
            column_index = col.column
            break

    if column_index is None:
        raise ValueError("Column 'QecolAlar' not found in the Excel sheet.")

    # Update the column with new values
    current_row = 3  # Start from row 3
    for year in range(years_sim):
        for month_index in range(12):
            sheet.cell(row=current_row, column=column_index).value = qecolAlar_values[month_index]
            current_row += 1

    # Save the workbook
    workbook.save("data.xlsx")
    upload_to_s3("data.xlsx", "mon-bucket-s3")

    # Rerun the Vensim model
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    variables_model = vensim_model.run(params={
        'INITIAL TIME': 1,
        'FINAL TIME': 12 * years_sim,
        'TIME STEP': 1
    })

    updated_outflow = variables_model['Sal Jucar']
    updated_deficit = variables_model['DéfQecolAlar']

    # Create updated graphs
    outflow_figure = go.Figure()
    outflow_figure.add_trace(go.Scatter(
        x=months, y=initial_outflow, mode="lines",
        name=f"Initial Outflow (QecoAlar = {initial_qecolAlar_value})",
        line=dict(dash="dot")
    ))
    outflow_figure.add_trace(go.Scatter(
        x=months, y=updated_outflow, mode="lines",
        name="Updated Outflow"
    ))
    outflow_figure.update_layout(
        title="Outflow Over Time",
        xaxis=dict(title="Time",tickmode="array", tickvals=tickvals, ticktext=ticktext,rangeslider=dict(visible=True),),  # Enable range slider for navigation
        yaxis=dict(title="hm³"),
        margin=dict(l=40, r=40, t=40, b=40),
        template="plotly_white",
    )

    deficit_figure = go.Figure()
    deficit_figure.add_trace(go.Scatter(
        x=months, y=initial_deficit, mode="lines",
        name=f"Initial Deficit (QecoAlar = {initial_qecolAlar_value})",
        line=dict(dash="dot")
    ))
    deficit_figure.add_trace(go.Scatter(
        x=months, y=updated_deficit, mode="lines",
        name="Updated Deficit"
    ))
    deficit_figure.update_layout(
        title="Deficit Over Time",
        xaxis=dict(title="Time",tickmode="array", tickvals=tickvals, ticktext=ticktext,rangeslider=dict(visible=True),),  # Enable range slider for navigation
        yaxis=dict(title="hm³"),
        margin=dict(l=40, r=40, t=40, b=40),
        template="plotly_white",
    )

    return outflow_figure, deficit_figure


#4.3 Population growth call back
#modal
@app.callback(
    Output("modal-population", "is_open"),
    [Input("info-icon-population", "n_clicks"), Input("close-population", "n_clicks")],
    [State("modal-population", "is_open")],
)
def toggle_modal_population(icon_click, close_click, is_open):
    if icon_click or close_click:
        return not is_open
    return is_open

#graphs
@app.callback(
    Output("demand-graph", "figure"),
    [Input("run-simulation", "n_clicks")],  # Button click to trigger
    [State("start-year", "value"), 
     State("end-year", "value"),
     State("variation-rate-input", "value")],
    prevent_initial_call=True
)

def update_population_graph(n_clicks, start_year, end_year, variation_rate):
    if n_clicks is None:
        raise dash.exceptions.PreventUpdate  # Prevent callback if no clicks
    # Calculate the number of months based on the selected years
    years_sim = end_year - start_year + 1
    months = np.arange(1, 12 * years_sim + 1)
    
    # Load and run the Vensim model
    vensim_model = pysd.load('WEFE Jucar (Simple).py')
    #change Variation and Activate 
    variables_model = vensim_model.run(params={'INITIAL TIME': 1,'FINAL TIME': 12 * years_sim,'TIME STEP': 1,'Variation Rate': variation_rate, '"Activar/Desactivar"': 1,})
    updated_urban_demand = variables_model['Total Demanda Urbana']
    
    # Generate labels for all months and October-only
    all_months_labels = []
    october_labels = []
    tickvals_october = []

    for year_index, year in enumerate(range(start_year, end_year + 1)):
        for month_index, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
            all_months_labels.append(f"{month} {year}")
            if month == "Oct":
                tickvals_october.append(year_index * 12 + month_index + 1)
                october_labels.append(f"Oct {year}")

    # Determine tick labels based on simulation length
    if years_sim > 3:
        tickvals = tickvals_october
        ticktext = october_labels
    else:
        tickvals = list(range(1, len(all_months_labels) + 1))
        ticktext = all_months_labels

    # Create a Plotly figure for urban demand
    urban_demand_figure = go.Figure()
    # initial urban demand trace
    urban_demand_figure.add_trace(
        go.Scatter(x=months,y=initial_urban_demand,mode="lines",name=f"Initial Urban Demand (Variation rate = {initial_variation_rate})",line=dict(dash="dot"),)
    )
    # updated urban demand trace
    urban_demand_figure.add_trace(
        go.Scatter( x=months,y=updated_urban_demand, mode="lines",name=f"Updated Urban Demand (Variation rate = {variation_rate})",)
    )
    # Customize the layout of the figure
    urban_demand_figure.update_layout(
        title="Urban Demand Over Time",
        xaxis=dict(title="Time", tickmode="array", tickvals=tickvals,  ticktext=ticktext, rangeslider=dict(visible=True), ),
        yaxis=dict(title="hm³"),
        margin=dict(l=40, r=40, t=40, b=40),
        template="plotly_white",
    )

    return urban_demand_figure

#5.3 divers callback
@app.callback(
    Output("divers-graph", "figure"),
    [Input("run-simulation", "n_clicks")],  # Button click to trigger
    [State("start-year", "value"), 
     State("end-year", "value"),
     State("selected", "value")],
    prevent_initial_call=True
)
def update_divers_graph(n_clicks, start_year, end_year, chosen_variable):
    if n_clicks is None or chosen_variable is None:
        raise dash.exceptions.PreventUpdate

    # Calculate the number of months based on the selected years
    years_sim = end_year - start_year + 1
    months = np.arange(1, 12 * years_sim + 1)

    # Simulate the Vensim model to calculate the chosen variable
    variable = variables_model_initial[chosen_variable]  

    # Generate labels for all months and October-only
    all_months_labels = []
    october_labels = []
    tickvals_october = []

    for year_index, year in enumerate(range(start_year, end_year + 1)):
        for month_index, month in enumerate(["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]):
            all_months_labels.append(f"{month} {year}")
            if month == "Oct":
                tickvals_october.append(year_index * 12 + month_index + 1)
                october_labels.append(f"Oct {year}")

    # Determine tick labels based on simulation length
    if years_sim > 3:
        tickvals = tickvals_october
        ticktext = october_labels
    else:
        tickvals = list(range(1, len(all_months_labels) + 1))
        ticktext = all_months_labels

    # Create the figure
    figure = go.Figure()
    figure.add_trace(
        go.Scatter(
            x=months,
            y=variable,
            mode="lines",
            name=f"Updated {chosen_variable}",
        )
    )

    # Update layout with dynamic tick values and labels
    figure.update_layout(
        title={
        "text": f"{chosen_variable} Over Time",
        "x": 0.5,  # Center the title
        "xanchor": "center",  # Ensure proper alignment
        "yanchor": "top",  # Keep it at the top
        "font": {"size": 20, "family": "Arial, sans-serif", "color": "black", "weight": "bold"}  # Bold and styling
        },
        xaxis=dict(
            title="Time",
            tickmode="array",
            tickvals=tickvals,
            ticktext=ticktext,
            rangeslider=dict(visible=True),  # Enable range slider for navigation
        ),
        yaxis=dict(title="hm³"),
        margin=dict(l=40, r=40, t=40, b=40),
        template="plotly_white",
    )

    return figure



# Run the app
if __name__ == "__main__":
    app.run_server(debug=True)
